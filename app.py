"""Application principale Flask — QA Bug Tracker & Matrice des Risques ISTQB."""

import os
import json
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
from config import config
from database import (
    init_db, get_db,
    create_matrix, get_all_matrices, get_matrix, delete_matrix,
    get_matrix_rows, add_matrix_row, update_matrix_row, delete_matrix_row,
)
from gitlab_client import fetch_issues, filter_bug_issues, search_project_by_name
from data_processor import (
    process_issues, parse_gitlab_csv, get_module_bug_counts,
    compute_probability_level, compute_risk,
    extract_modules, extract_fonctionnalite,
    PROBABILITY_LEVELS, IMPACT_LEVELS, RISK_LABELS,
)

app = Flask(__name__)
app.secret_key = config.SECRET_KEY


# ─── Cache simple en mémoire ────────────────────────────────────
_issues_cache = {"data": None, "timestamp": None}


def get_cached_issues():
    """Récupérer les issues avec cache de 30 minutes."""
    now = datetime.now()
    if (_issues_cache["data"] is not None and
            _issues_cache["timestamp"] and
            (now - _issues_cache["timestamp"]).seconds < 1800):
        return _issues_cache["data"]

    issues = fetch_issues()
    bug_issues = filter_bug_issues(issues)
    _issues_cache["data"] = bug_issues
    _issues_cache["timestamp"] = now
    return bug_issues


# ─── Pages ───────────────────────────────────────────────────────

@app.route("/")
def index():
    """Page d'accueil → Dashboard bugs."""
    return render_template("dashboard.html",
                           is_mock=config.is_mock_mode,
                           current_year=datetime.now().year)


@app.route("/risk-matrix")
def risk_matrix_page():
    """Page matrice des risques."""
    matrices = get_all_matrices()
    return render_template("risk_matrix.html",
                           matrices=matrices,
                           is_mock=config.is_mock_mode,
                           impact_levels=IMPACT_LEVELS,
                           probability_levels=PROBABILITY_LEVELS)


# ─── API Dashboard ───────────────────────────────────────────────

@app.route("/api/bugs-data")
def api_bugs_data():
    """Données des bugs pour le dashboard."""
    year = request.args.get("year")
    issues = get_cached_issues()
    result = process_issues(issues, year=year)
    return jsonify(result)


@app.route("/api/search-project")
def api_search_project():
    """Rechercher un projet GitLab par nom."""
    name = request.args.get("name", "")
    if not name:
        return jsonify([])
    projects = search_project_by_name(name)
    return jsonify([{"id": p["id"], "name": p.get("name_with_namespace", p.get("name", ""))}
                    for p in projects])


@app.route("/api/refresh-cache", methods=["POST"])
def api_refresh_cache():
    """Forcer le rafraîchissement du cache."""
    _issues_cache["data"] = None
    _issues_cache["timestamp"] = None
    get_cached_issues()
    return jsonify({"status": "ok"})


# ─── API Matrices ────────────────────────────────────────────────

@app.route("/api/matrices", methods=["GET"])
def api_list_matrices():
    """Lister toutes les matrices."""
    matrices = get_all_matrices()
    for m in matrices:
        rows = get_matrix_rows(m["id"])
        m["row_count"] = len(rows)
    return jsonify(matrices)


@app.route("/api/matrices", methods=["POST"])
def api_create_matrix():
    """Créer une nouvelle matrice."""
    data = request.json
    name = data.get("name", "").strip()
    version = data.get("version", "").strip()
    parent_id = data.get("parent_id")

    if not name or not version:
        return jsonify({"error": "Nom et version requis"}), 400

    matrix_id = create_matrix(name, version, parent_id=parent_id)
    return jsonify({"id": matrix_id, "status": "created"})


@app.route("/api/matrices/<int:matrix_id>", methods=["GET"])
def api_get_matrix(matrix_id):
    """Récupérer une matrice avec ses lignes et les probabilités calculées."""
    matrix = get_matrix(matrix_id)
    if not matrix:
        return jsonify({"error": "Matrice non trouvée"}), 404

    rows = get_matrix_rows(matrix_id)

    # Calculer les probabilités basées sur les données de bugs
    issues = get_cached_issues()
    module_bug_counts = get_module_bug_counts(issues)

    enriched_rows = []
    for row in rows:
        prob_level = compute_probability_level(row["module"], module_bug_counts)
        impact_level_data = next(
            (il for il in IMPACT_LEVELS if il["key"] == row["impact_level"]),
            IMPACT_LEVELS[0]
        )
        risk_value, risk_info = compute_risk(prob_level["value"], impact_level_data["value"])

        enriched_rows.append({
            **row,
            "probability": prob_level,
            "impact": impact_level_data,
            "risk_value": risk_value,
            "risk_label": risk_info["label"],
            "risk_color": risk_info["color"],
        })

    return jsonify({
        "matrix": matrix,
        "rows": enriched_rows,
        "module_bug_counts": module_bug_counts,
    })


@app.route("/api/matrices/<int:matrix_id>", methods=["DELETE"])
def api_delete_matrix(matrix_id):
    """Supprimer une matrice."""
    delete_matrix(matrix_id)
    return jsonify({"status": "deleted"})


# ─── API Lignes de matrice ───────────────────────────────────────

@app.route("/api/matrices/<int:matrix_id>/rows", methods=["POST"])
def api_add_row(matrix_id):
    """Ajouter une ligne manuellement."""
    data = request.json
    row_id = add_matrix_row(
        matrix_id,
        module=data.get("module", "").upper(),
        fonctionnalite=data.get("fonctionnalite", ""),
        gitlab_iid=data.get("gitlab_iid", ""),
        impact_level=data.get("impact_level", "non_defini"),
        is_manual=1,
    )
    return jsonify({"id": row_id, "status": "added"})


@app.route("/api/matrix-rows/<int:row_id>", methods=["PUT"])
def api_update_row(row_id):
    """Mettre à jour une ligne."""
    data = request.json
    update_matrix_row(row_id, **data)
    return jsonify({"status": "updated"})


@app.route("/api/matrix-rows/<int:row_id>", methods=["DELETE"])
def api_delete_row(row_id):
    """Supprimer une ligne."""
    delete_matrix_row(row_id)
    return jsonify({"status": "deleted"})


# ─── Import CSV ──────────────────────────────────────────────────

@app.route("/api/matrices/<int:matrix_id>/import-csv", methods=["POST"])
def api_import_csv(matrix_id):
    """Importer un CSV GitLab dans une matrice."""
    matrix = get_matrix(matrix_id)
    if not matrix:
        return jsonify({"error": "Matrice non trouvée"}), 404

    if "file" not in request.files:
        return jsonify({"error": "Aucun fichier fourni"}), 400

    file = request.files["file"]
    if not file.filename.endswith(".csv"):
        return jsonify({"error": "Le fichier doit être au format CSV"}), 400

    csv_content = file.read()
    parsed_rows = parse_gitlab_csv(csv_content)

    # Récupérer les bug counts pour calculer les probabilités
    issues = get_cached_issues()
    module_bug_counts = get_module_bug_counts(issues)

    imported = 0
    for row in parsed_rows:
        # Seuls les tickets bug prod/préprod
        if not row["category"]:
            continue

        module_display = ", ".join(row["modules"])
        prob_level = compute_probability_level(module_display, module_bug_counts)

        add_matrix_row(
            matrix_id,
            module=module_display,
            fonctionnalite=row["fonctionnalite"],
            gitlab_iid=row["gitlab_iid"],
            probability_level=prob_level["key"],
            impact_level="non_defini",
            is_manual=0,
        )
        imported += 1

    return jsonify({"status": "imported", "count": imported})


# ─── Export Excel ────────────────────────────────────────────────

@app.route("/api/matrices/<int:matrix_id>/export-excel")
def api_export_excel(matrix_id):
    """Exporter une matrice en Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import tempfile

    matrix = get_matrix(matrix_id)
    if not matrix:
        return jsonify({"error": "Matrice non trouvée"}), 404

    rows = get_matrix_rows(matrix_id)
    issues = get_cached_issues()
    module_bug_counts = get_module_bug_counts(issues)

    wb = Workbook()

    # ── Feuille 1 : Détail par ticket ──
    ws1 = wb.active
    ws1.title = "Détail par ticket"

    headers = ["Module", "Fonctionnalité", "N° GitLab", "Probabilité", "Impact", "Risque"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    risk_colors = {
        "Faible": "92D050",
        "Modéré": "FFC000",
        "Élevé": "FF6600",
        "Critique": "FF0000",
        "Non défini": "D9D9D9",
    }

    for i, row in enumerate(rows, 2):
        prob_level = compute_probability_level(row["module"], module_bug_counts)
        impact_data = next(
            (il for il in IMPACT_LEVELS if il["key"] == row["impact_level"]),
            IMPACT_LEVELS[0]
        )
        risk_value, risk_info = compute_risk(prob_level["value"], impact_data["value"])

        ws1.cell(row=i, column=1, value=row["module"]).border = thin_border
        ws1.cell(row=i, column=2, value=row["fonctionnalite"]).border = thin_border
        ws1.cell(row=i, column=3, value=row["gitlab_iid"]).border = thin_border
        ws1.cell(row=i, column=4, value=prob_level["label"]).border = thin_border
        ws1.cell(row=i, column=5, value=impact_data["label"]).border = thin_border

        risk_cell = ws1.cell(row=i, column=6, value=risk_info["label"])
        risk_cell.border = thin_border
        color_hex = risk_colors.get(risk_info["label"], "D9D9D9")
        risk_cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        if risk_info["label"] in ("Élevé", "Critique"):
            risk_cell.font = Font(color="FFFFFF", bold=True)

    # Ajuster les largeurs
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 45
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 15
    ws1.column_dimensions["E"].width = 12
    ws1.column_dimensions["F"].width = 12

    # ── Feuille 2 : Matrice des risques (grille) ──
    ws2 = wb.create_sheet("Matrice des risques")

    ws2.cell(row=1, column=1, value="Probabilité \\ Impact")
    impact_headers = ["Mineur", "Modéré", "Majeur", "Critique"]
    prob_headers = ["Très faible", "Faible", "Moyenne", "Élevée", "Très élevée"]

    for col, imp in enumerate(impact_headers, 2):
        cell = ws2.cell(row=1, column=col, value=imp)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    grid_colors = {1: "92D050", 2: "FFC000", 3: "FF6600", 4: "FF0000"}

    for row_idx, prob_label in enumerate(prob_headers, 2):
        cell = ws2.cell(row=row_idx, column=1, value=prob_label)
        cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = thin_border

        for col_idx, imp_val in enumerate(range(1, 5), 2):
            prob_val = row_idx - 1
            risk_val, risk_info = compute_risk(prob_val, imp_val)
            cell = ws2.cell(row=row_idx, column=col_idx, value=risk_info["label"])
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            color = grid_colors.get(risk_val, "D9D9D9")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            if risk_val >= 3:
                cell.font = Font(color="FFFFFF", bold=True)

    ws2.column_dimensions["A"].width = 18
    for col_letter in ["B", "C", "D", "E"]:
        ws2.column_dimensions[col_letter].width = 14

    # Sauvegarder
    filename = f"matrice_risques_{matrix['version'].replace(' ', '_')}.xlsx"
    filepath = os.path.join(tempfile.gettempdir(), filename)
    wb.save(filepath)

    return send_file(filepath, as_attachment=True, download_name=filename)


# ─── Initialisation ──────────────────────────────────────────────

init_db()

if __name__ == "__main__":
    app.run(
        host=config.APP_HOST,
        port=config.APP_PORT,
        debug=config.is_local,
    )
