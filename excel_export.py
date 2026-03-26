"""
Export de la matrice des risques en fichier Excel formaté (par ticket).
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── Styles ───────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT  = Font(bold=True, size=14, color="1A237E")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

RISK_FILLS = {
    "critique":   PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid"),
    "eleve":      PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid"),
    "moyen":      PatternFill(start_color="F1C40F", end_color="F1C40F", fill_type="solid"),
    "faible":     PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid"),
    "non_evalue": PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid"),
}

IMPACT_FILLS = {
    "critique":   PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid"),
    "majeur":     PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid"),
    "modere":     PatternFill(start_color="F1C40F", end_color="F1C40F", fill_type="solid"),
    "mineur":     PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid"),
    "non_defini": PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid"),
}


def _cell(ws, row, col, value="", fill=None, font=None, align="center", wrap=False, border=True):
    c = ws.cell(row=row, column=col, value=value)
    if fill:   c.fill = fill
    if font:   c.font = font
    if border: c.border = THIN_BORDER
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    return c


def export_risk_matrix_to_excel(matrix_info, rows):
    """
    Génère un fichier Excel depuis la liste de lignes enrichies.
    rows : liste de dicts avec module, fonctionnalite, gitlab_iid,
           probability (dict), impact (dict), risk_label, risk_key, risk_color.
    Retourne un BytesIO buffer.
    """
    wb = Workbook()

    # ── Feuille 1 : Détail par ticket ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Matrice des Risques"
    ws.freeze_panes = "A6"

    # En-tête du rapport
    ws.merge_cells("A1:G1")
    _cell(ws, 1, 1,
          f"Matrice des Risques ISTQB — {matrix_info.get('name', '')} ({matrix_info.get('version', '')})",
          font=TITLE_FONT, border=False)

    ws.merge_cells("A2:G2")
    _cell(ws, 2, 1,
          f"Créée le {matrix_info.get('created_at', '')[:16]}",
          font=Font(italic=True, color="666666"), border=False)

    # Ligne vide
    ws.append([])
    ws.append([])

    # En-têtes colonnes
    headers = ["Module", "Fonctionnalité", "N° GitLab", "Probabilité", "Impact", "Risque", "Commentaire"]
    for col, h in enumerate(headers, 1):
        _cell(ws, 5, col, h, fill=HEADER_FILL, font=HEADER_FONT)

    # Données
    for i, row in enumerate(rows, 6):
        prob  = row.get("probability") or {}
        impact = row.get("impact") or {}

        _cell(ws, i, 1, row.get("module", ""),          align="left")
        _cell(ws, i, 2, row.get("fonctionnalite", ""),  align="left", wrap=True)
        _cell(ws, i, 3, row.get("gitlab_iid", ""))
        _cell(ws, i, 4, prob.get("label",  "N/A"))
        _cell(ws, i, 5, impact.get("label","N/A"))
        _cell(ws, i, 6, row.get("risk_label", "N/A"))
        _cell(ws, i, 7, row.get("comment", ""),         align="left", wrap=True)

        # Couleur impact
        impact_fill = IMPACT_FILLS.get(row.get("impact_level", "non_defini"))
        if impact_fill:
            ws.cell(row=i, column=5).fill = impact_fill
            if row.get("impact_level") in ("critique", "majeur"):
                ws.cell(row=i, column=5).font = Font(color="FFFFFF", bold=True)

        # Couleur risque
        risk_fill = RISK_FILLS.get(row.get("risk_key", "non_evalue"))
        if risk_fill:
            ws.cell(row=i, column=6).fill = risk_fill
            if row.get("risk_key") in ("critique", "eleve"):
                ws.cell(row=i, column=6).font = Font(color="FFFFFF", bold=True)

    # Largeurs
    for col, width in enumerate([20, 45, 12, 16, 12, 12, 30], 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[5].height = 22

    # ── Feuille 2 : Grille visuelle ───────────────────────────────────────────
    ws2 = wb.create_sheet("Grille Visuelle")

    # Matrice de risque ISTQB 5x4
    prob_labels   = ["Très élevée", "Élevée", "Moyenne", "Faible", "Très faible"]
    impact_labels = ["Mineur", "Modéré", "Majeur", "Critique"]
    risk_matrix   = {
        5: {1: "moyen",    2: "eleve",    3: "critique", 4: "critique"},
        4: {1: "moyen",    2: "eleve",    3: "eleve",    4: "critique"},
        3: {1: "faible",   2: "moyen",    3: "eleve",    4: "eleve"},
        2: {1: "faible",   2: "faible",   3: "moyen",    4: "moyen"},
        1: {1: "faible",   2: "faible",   3: "faible",   4: "moyen"},
    }
    risk_labels_map = {
        "critique": "Critique", "eleve": "Élevé", "moyen": "Moyen", "faible": "Faible"
    }

    ws2["A1"] = "Grille des Risques ISTQB"
    ws2["A1"].font = TITLE_FONT

    # Coin + headers impact
    _cell(ws2, 3, 1, "Prob \\ Impact", fill=PatternFill(start_color="37474F", end_color="37474F", fill_type="solid"),
          font=Font(color="FFFFFF", bold=True))
    for col, label in enumerate(impact_labels, 2):
        _cell(ws2, 3, col, label, fill=HEADER_FILL, font=HEADER_FONT)

    # Lignes probabilité
    for row_i, (prob_label, prob_val) in enumerate(zip(prob_labels, [5,4,3,2,1]), 4):
        _cell(ws2, row_i, 1, prob_label,
              fill=PatternFill(start_color="37474F", end_color="37474F", fill_type="solid"),
              font=Font(color="FFFFFF", bold=True), align="right")
        for col_i, imp_val in enumerate([1,2,3,4], 2):
            risk_key = risk_matrix[prob_val][imp_val]
            _cell(ws2, row_i, col_i,
                  risk_labels_map[risk_key],
                  fill=RISK_FILLS[risk_key],
                  font=Font(color="FFFFFF" if risk_key in ("critique","eleve") else "333333", bold=True))

    # Légende modules par cellule
    ws2["A11"] = "Modules par niveau de risque (matrice courante)"
    ws2["A11"].font = Font(bold=True)

    # Regrouper les modules par (prob_value, impact_value)
    from collections import defaultdict
    cell_modules = defaultdict(list)
    for row in rows:
        p = (row.get("probability") or {}).get("value", 0)
        iv = (row.get("impact") or {}).get("value", 0)
        if p and iv:
            cell_modules[(p, iv)].append(row.get("module", ""))

    row_start = 13
    for (pv, iv), mods in sorted(cell_modules.items(), reverse=True):
        prob_label_  = next((l for k,l,v,_ in [("tres_elevee","Très élevée",5,20),("elevee","Élevée",4,12),
                              ("moyenne","Moyenne",3,6),("faible","Faible",2,2),("tres_faible","Très faible",1,0)] if v==pv), "?")
        impact_label_ = ["","Mineur","Modéré","Majeur","Critique"][iv]
        ws2.cell(row=row_start, column=1, value=f"{prob_label_} x {impact_label_} → {', '.join(set(mods))}")
        row_start += 1

    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 22

    # ── Feuille 3 : Légende ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Légende ISTQB")

    ws3["A1"] = "Échelle de Probabilité"
    ws3["A1"].font = Font(bold=True, size=12)

    for col, h in enumerate(["Niveau", "Seuil (%)", "Description"], 1):
        _cell(ws3, 3, col, h, fill=HEADER_FILL, font=HEADER_FONT)

    prob_data = [
        ("Très élevée", "≥ 20%", "Module concentrant une part très importante des bugs"),
        ("Élevée",      "≥ 12%", "Module avec une fréquence élevée de bugs"),
        ("Moyenne",     "≥ 6%",  "Module avec une fréquence modérée de bugs"),
        ("Faible",      "≥ 2%",  "Module avec peu de bugs"),
        ("Très faible", "< 2%",  "Module rarement affecté par les bugs"),
    ]
    for ri, row_data in enumerate(prob_data, 4):
        for ci, v in enumerate(row_data, 1):
            _cell(ws3, ri, ci, v)

    ws3["A10"] = "Échelle d'Impact (saisie manuelle)"
    ws3["A10"].font = Font(bold=True, size=12)

    for col, h in enumerate(["Niveau", "Description"], 1):
        _cell(ws3, 12, col, h, fill=HEADER_FILL, font=HEADER_FONT)

    impact_data = [
        ("Critique", "Impact business majeur — perte de données, indisponibilité totale"),
        ("Majeur",   "Fonctionnalité clé inutilisable, contournement difficile"),
        ("Modéré",   "Fonctionnalité dégradée, contournement possible"),
        ("Mineur",   "Gêne mineure, impact cosmétique ou ergonomique"),
    ]
    for ri, (level, desc) in enumerate(impact_data, 13):
        _cell(ws3, ri, 1, level)
        _cell(ws3, ri, 2, desc, align="left")

    for col in range(1, 4):
        ws3.column_dimensions[get_column_letter(col)].width = 30

    # ── Export ────────────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
